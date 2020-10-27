import tkinter
import tkinter.messagebox
import tkinter.filedialog
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from string import Template
import re


#  define main function
def main():
    # 创建导出目录
    if not os.path.exists('生成'):
        os.mkdir('生成')
    source_path = ""

    def get_source_path():
        nonlocal source_path
        source_path = tkinter.filedialog.askopenfilename(title='选择文件', filetypes=[('表格', '.xlsx')])
        label.config(text='文件路径：' + source_path)

    # create main top window
    top = tkinter.Tk()
    # set window-size
    top.geometry('640x360')
    # set window-title
    top.title("excel模版转换")
    # create text-obj
    label = tkinter.Label(top, text="文件路径:", font='Arial-32', fg='red')
    label.pack(side='top')

    # create panel container for btn
    panel = tkinter.Frame(top)
    # create btn-obj set into certain panel
    btn1 = tkinter.Button(panel, text='打开数据源文件', command=get_source_path)
    btn1.pack(side='left')
    btn2 = tkinter.Button(panel, text='生成', command=lambda: read_cell_maps(source_path))
    btn2.pack(side='right')
    panel.pack(side='bottom')
    # start main loop
    tkinter.mainloop()


# read excel-map text
def read_cell_maps(source_path):
    # read source-xls
    if os.path.isfile(source_path):
        source_workbook = load_workbook(source_path)
    else:
        tkinter.messagebox.showinfo('温馨提示', '源文件不存在')
        return

    # read target-model-xls
    if os.path.isfile('./模板.xlsx'):
        model_workbook = load_workbook('./模板.xlsx')
    else:
        tkinter.messagebox.showinfo('温馨提示', '模板文件不存在')
        return

    top1 = '上海明汯投资管理有限公司-明汯价值成长1期私募投资基金'
    top2 = '华泰金融控股(香港)有限公司-自有资金'
    result = []
    for sheet in source_workbook.worksheets:
        rowcount = sheet.max_row
        for row in sheet.rows:
            rowarray = row[2].value.split(',')
            if len(rowarray) < 3:
                continue
            if top1 in row[2].value.split(',') and top2 in row[2].value.split(','):
                result.append(row)

    # 另存为
    for data in result:
        print(data[0].value,data[1].value)
    model_workbook.save('./生成/{}.xlsx'.format('十大'))
    tkinter.messagebox.showinfo('温馨提示', '转换完成！')


if __name__ == '__main__':
    main()
